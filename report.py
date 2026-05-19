import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use("Agg")  # non-interactive backend for saving files
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
import warnings
warnings.filterwarnings("ignore")

# ------------------------------------------------------------
# LOAD DATA
# ------------------------------------------------------------
df = pd.read_csv("clean_reviews.csv")
df["date"] = pd.to_datetime(df["date"])

# ------------------------------------------------------------
# RECREATE ALL FINDINGS
# ------------------------------------------------------------

# Finding 1 — ratings
avg_rating = df.groupby("platform")["rating"].mean().round(2)

# Finding 2 — sentiment distribution
sentiment = df.groupby(["platform", "sentiment"]).size().unstack(fill_value=0)
sentiment_pct = (sentiment.div(sentiment.sum(axis=1), axis=0) * 100).round(1)

# Finding 2b — complaint categories
keywords = {
    "Late delivery"    : ["late", "delay", "slow", "hours", "waiting", "wait"],
    "Order cancelled"  : ["cancel", "cancelled", "cancellation"],
    "Wrong item"       : ["wrong", "missing", "not delivered", "incomplete"],
    "Refund issue"     : ["refund", "money", "charged", "payment"],
    "App issue"        : ["app", "crash", "bug", "error", "glitch"]
}
bad = df[df["rating"] <= 2].copy()
for category, words in keywords.items():
    bad[category] = bad["review"].str.lower().str.contains("|".join(words), na=False)
complaint_counts = bad.groupby("platform")[list(keywords.keys())].sum()

# Finding 3 — review length by sentiment
length_sentiment = df.groupby(["platform", "sentiment"])["review_length"].mean().round(0).unstack()

# Finding 4 — ops health score
positive_score = sentiment_pct["positive"]
neg_length = df[df["sentiment"] == "negative"].groupby("platform")["review_length"].mean()
max_len = neg_length.max()
complaint_rate = (complaint_counts.sum(axis=1) / bad.groupby("platform").size() * 100)
scores = pd.DataFrame({
    "Rating score (40%)"    : ((avg_rating - 1) / 4 * 100).round(1),
    "Positive score (30%)"  : positive_score.round(1),
    "Complaint score (20%)" : (100 - complaint_rate).clip(lower=0).round(1),
    "Length score (10%)"    : (100 - (neg_length / max_len * 100)).round(1)
})
scores["Ops Health Score"] = (
    scores["Rating score (40%)"]    * 0.40 +
    scores["Positive score (30%)"]  * 0.30 +
    scores["Complaint score (20%)"] * 0.20 +
    scores["Length score (10%)"]    * 0.10
).round(1)

# ------------------------------------------------------------
# GENERATE CHARTS
# ------------------------------------------------------------
colors = {"Blinkit": "#F4C430", "Swiggy": "#FC8019", "Zepto": "#8B5CF6"}
platforms = ["Blinkit", "Swiggy", "Zepto"]
bar_colors = [colors[p] for p in platforms]

# Chart 1 — avg rating
fig, ax = plt.subplots(figsize=(7, 4))
bars = ax.bar(platforms, avg_rating[platforms], color=bar_colors, width=0.5)
ax.set_title("Average Rating by Platform", fontsize=13, fontweight="bold")
ax.set_ylabel("Rating (out of 5)")
ax.set_ylim(0, 5)
for bar, val in zip(bars, avg_rating[platforms]):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1,
            str(val), ha="center", fontsize=11, fontweight="bold")
plt.tight_layout()
plt.savefig("chart1_ratings.png", dpi=150)
plt.close()
print("Chart 1 saved")

# Chart 2 — sentiment distribution
fig, ax = plt.subplots(figsize=(7, 4))
x = range(len(platforms))
width = 0.25
ax.bar([i - width for i in x], sentiment_pct.loc[platforms, "positive"],
       width=width, label="Positive", color="#22C55E")
ax.bar(x, sentiment_pct.loc[platforms, "neutral"],
       width=width, label="Neutral", color="#94A3B8")
ax.bar([i + width for i in x], sentiment_pct.loc[platforms, "negative"],
       width=width, label="Negative", color="#EF4444")
ax.set_title("Sentiment Distribution by Platform (%)", fontsize=13, fontweight="bold")
ax.set_ylabel("Percentage of reviews")
ax.set_xticks(list(x))
ax.set_xticklabels(platforms)
ax.legend()
plt.tight_layout()
plt.savefig("chart2_sentiment.png", dpi=150)
plt.close()
print("Chart 2 saved")

# Chart 3 — complaint categories
fig, ax = plt.subplots(figsize=(10, 6))
complaint_data = complaint_counts.loc[platforms]
x = range(len(keywords))
width = 0.25

for i, platform in enumerate(platforms):
    offset = (i - 1) * width
    vals = [complaint_data.loc[platform, cat] for cat in keywords.keys()]
    bars = ax.bar([xi + offset for xi in x], vals, width=width,
                  label=platform, color=bar_colors[i])
    for bar, val in zip(bars, vals):
        if val > 0:
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.2,
                    str(int(val)), ha="center", va="bottom", fontsize=7)

ax.set_title("Complaint Categories by Platform", fontsize=14, fontweight="bold", pad=15)
ax.set_ylabel("Number of mentions", fontsize=11)
ax.set_xticks(list(x))
ax.set_xticklabels(list(keywords.keys()), fontsize=11)
ax.legend(fontsize=11)
ax.set_ylim(0, complaint_data.values.max() + 15)
ax.grid(axis="y", alpha=0.3)
plt.tight_layout()
plt.savefig("chart3_complaints.png", dpi=150)
plt.close()
print("Chart 3 saved")

# Chart 4 — ops health score
fig, ax = plt.subplots(figsize=(7, 4))
score_vals = scores.loc[platforms, "Ops Health Score"]
bars = ax.bar(platforms, score_vals, color=bar_colors, width=0.5)
ax.set_title("Ops Health Score by Platform (0-100)", fontsize=13, fontweight="bold")
ax.set_ylabel("Score")
ax.set_ylim(0, 100)
for bar, val in zip(bars, score_vals):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 1,
            str(val), ha="center", fontsize=11, fontweight="bold")
plt.tight_layout()
plt.savefig("chart4_ops_score.png", dpi=150)
plt.close()
print("Chart 4 saved")

# ------------------------------------------------------------
# GENERATE EXCEL REPORT
# ------------------------------------------------------------
wb = Workbook()

# --- Sheet 1: Summary ---
ws1 = wb.active
ws1.title = "Summary"

header_font = Font(bold=True, size=12)
header_fill = PatternFill("solid", fgColor="1E3A5F")
header_font_white = Font(bold=True, size=12, color="FFFFFF")

ws1["A1"] = "Quick Commerce Ops Intelligence Report"
ws1["A1"].font = Font(bold=True, size=16)
ws1["A2"] = "Data source: Google Play Store reviews (1,500 reviews) + Google Trends"
ws1["A3"] = "Platforms analyzed: Blinkit, Swiggy, Zepto"
ws1["A4"] = "Analysis period: May 2026"

ws1.append([])
ws1.append(["Ops Health Score Ranking"])
ws1["A6"].font = header_font

for i, platform in enumerate(["Blinkit", "Swiggy", "Zepto"]):
    score = scores.loc[platform, "Ops Health Score"]
    ws1.append([f"{i+1}. {platform}", f"{score}/100"])

ws1.append([])
ws1.append(["Key Findings"])
ws1["A10"].font = header_font
ws1.append(["1", "Zepto has the highest 1-star review rate at 41.6% vs Blinkit's 12.6%"])
ws1.append(["2", "Zepto leads complaint volume across app issues (83), late delivery (38), refund problems (31)"])
ws1.append(["3", "Swiggy leads on order cancellations with 33 mentions vs Blinkit's 4"])
ws1.append(["4", "Negative reviews are 4-7x longer than positive reviews across all platforms"])
ws1.append(["5", "Blinkit dominates Google search interest at 43.8 avg vs Zepto 25 and Swiggy Instamart 2.4"])

ws1.column_dimensions["A"].width = 12
ws1.column_dimensions["B"].width = 80

# --- Sheet 2: Ratings ---
ws2 = wb.create_sheet("Ratings")
ws2.append(["Platform", "Avg Rating", "1-Star %", "5-Star %"])
ws2["A1"].font = header_font
for platform in platforms:
    one_star_pct = round(sentiment_pct.loc[platform, "negative"], 1)
    five_star_pct = round(sentiment_pct.loc[platform, "positive"], 1)
    ws2.append([platform, float(avg_rating[platform]), one_star_pct, five_star_pct])
ws2.column_dimensions["A"].width = 15
ws2.column_dimensions["B"].width = 15
ws2.column_dimensions["C"].width = 15
ws2.column_dimensions["D"].width = 15

# --- Sheet 3: Complaint Categories ---
ws3 = wb.create_sheet("Complaint Categories")
complaint_export = complaint_counts.loc[platforms].reset_index()
for r in dataframe_to_rows(complaint_export, index=False, header=True):
    ws3.append(r)
for cell in ws3[1]:
    cell.font = header_font
ws3.column_dimensions["A"].width = 15

# --- Sheet 4: Ops Health Score ---
ws4 = wb.create_sheet("Ops Health Score")
score_export = scores.loc[platforms].reset_index()
for r in dataframe_to_rows(score_export, index=False, header=True):
    ws4.append(r)
for cell in ws4[1]:
    cell.font = header_font
for col in ["A","B","C","D","E","F"]:
    ws4.column_dimensions[col].width = 22


print("Files: chart3_ratings.png")